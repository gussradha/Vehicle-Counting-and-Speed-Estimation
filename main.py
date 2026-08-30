"""
Vehicle Counting and Speed Estimation
======================================

Counts vehicles (LV/HV/MC) and estimates their speed from traffic/CCTV video
using YOLO detection + tracking, writing results to an Excel file broken down
by configurable time intervals (e.g. every 5, 15, or 30 minutes).

Usage examples:
    python main.py --video data/sample.mp4 --interval 15
    python main.py --video data/sample.mp4 --interval 5 --realtime
    python main.py --video data/sample.mp4 --model yolov10s.pt --device mps

Run `python main.py --help` for the full list of options.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import time
from dataclasses import dataclass, field

import cv2
import cvzone
import numpy as np
import torch
from openpyxl import Workbook, load_workbook
from ultralytics import YOLO

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# Directory this script lives in, used to resolve relative paths (coco.txt, etc.)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Region of Interest (ROI) where vehicles are counted, in pixel coordinates
# after the frame has been resized to FRAME_SIZE. Adjust these points to match
# your own camera/video (e.g. using tools/pick_roi.py).
DEFAULT_ROI = [(603, 398), (690, 398), (670, 318), (600, 315)]
FRAME_SIZE = (1020, 500)

# Pixel-distance-to-speed calibration factor. This value depends on the
# camera's angle and distance from the road, so it MUST be recalibrated for
# any other video (e.g. by measuring a real-world distance between two points
# on the road and comparing it to the pixel distance in the frame).
PIXEL_TO_METER_SCALE = 30
SPEED_MIN_KMH, SPEED_MAX_KMH = 5, 80

VEHICLE_CLASSES = ("car", "bus", "truck", "motorcycle")


@dataclass
class IntervalCounts:
    """Accumulated counts and speed records for a single time interval."""

    car_ids: set = field(default_factory=set)
    bus_ids: set = field(default_factory=set)
    truck_ids: set = field(default_factory=set)
    motorcycle_ids: set = field(default_factory=set)
    speed_records: list = field(default_factory=list)  # (track_id, label, speed_kmh)
    speed_recorded_ids: set = field(default_factory=set)

    @property
    def lv(self) -> int:
        """Light Vehicle count (cars)."""
        return len(self.car_ids)

    @property
    def hv(self) -> int:
        """Heavy Vehicle count (trucks + buses)."""
        return len(self.truck_ids) + len(self.bus_ids)

    @property
    def mc(self) -> int:
        """Motorcycle count."""
        return len(self.motorcycle_ids)

    def is_empty(self) -> bool:
        return not (self.car_ids or self.bus_ids or self.truck_ids or self.motorcycle_ids)


def format_hms(total_seconds: float) -> str:
    """Format a number of seconds as an HH:MM:SS string."""
    h = int(total_seconds // 3600)
    m = int((total_seconds % 3600) // 60)
    s = int(total_seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def resolve_device(preferred: str | None) -> str:
    """Pick the inference device: use the user's preference if given, else auto-detect."""
    if preferred:
        return preferred
    if torch.cuda.is_available():
        return "cuda"
    if torch.backends.mps.is_available():
        return "mps"
    return "cpu"


def load_class_names(path: str) -> list[str]:
    with open(path, "r") as f:
        return f.read().splitlines()


def load_roi(roi_file: str | None) -> list[tuple[int, int]]:
    """Load ROI points from a JSON file (produced by tools/pick_roi.py), or fall back to DEFAULT_ROI."""
    if not roi_file:
        logger.info("No --roi-file provided, using the built-in DEFAULT_ROI.")
        return DEFAULT_ROI
    with open(roi_file, "r") as f:
        points = json.load(f)
    roi = [tuple(p) for p in points]
    logger.info("ROI loaded from %s: %s", roi_file, roi)
    return roi


def save_interval_to_excel(
    output_path: str,
    counts: IntervalCounts,
    start_seconds: float,
    end_seconds: float,
) -> None:
    """Append one row of interval counts + its speed records to the Excel output file."""
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Survey Data" in wb.sheetnames:
        sheet = wb["Survey Data"]
    else:
        sheet = wb.create_sheet("Survey Data")
        sheet.append(["No", "Time Interval", "LV", "HV", "MC"])

    row_number = sheet.max_row  # row 1 is the header
    interval_label = f"{format_hms(start_seconds)} - {format_hms(end_seconds)}"
    sheet.append([row_number, interval_label, counts.lv, counts.hv, counts.mc])

    if "Speed" in wb.sheetnames:
        sheet_speed = wb["Speed"]
    else:
        sheet_speed = wb.create_sheet("Speed")
        sheet_speed.append(["ID", "Vehicle Type", "Speed (km/h)", "Time Interval"])

    for track_id, label, speed_kmh in counts.speed_records:
        sheet_speed.append([track_id, label, f"{speed_kmh:.2f}", interval_label])

    for ws in (sheet, sheet_speed):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_path)
    logger.info(
        "Interval %s saved -> LV:%d HV:%d MC:%d (%d speed records)",
        interval_label, counts.lv, counts.hv, counts.mc, len(counts.speed_records),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Count vehicles and estimate their speed from a video, saved per interval to Excel.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--video", default="data/sample.mp4", help="Path to the source video")
    parser.add_argument("--output", default="output.xlsx", help="Path to the Excel output file")
    parser.add_argument("--model", default="yolov10x.pt", help="Path/name of the YOLO weights to use")
    parser.add_argument(
        "--classes", default=os.path.join(BASE_DIR, "coco.txt"), help="Path to the class names file (coco.txt)"
    )
    parser.add_argument(
        "--interval", type=float, default=15, help="Excel logging interval, in minutes"
    )
    parser.add_argument(
        "--device", default=None, choices=["cuda", "mps", "cpu"],
        help="Force a specific device; default: auto-detect (CUDA > MPS > CPU)",
    )
    parser.add_argument(
        "--realtime", action="store_true",
        help="Sync playback to the video's real speed (frames may be skipped while catching up). "
             "Good for quick previews; NOT recommended for final data collection.",
    )
    parser.add_argument(
        "--no-display", action="store_true", help="Run without showing the video window (faster, headless)"
    )
    parser.add_argument(
        "--roi-file", default=None,
        help="Path to a JSON file with ROI points (produced by tools/pick_roi.py). "
             "Default: use the DEFAULT_ROI defined in source code.",
    )
    parser.add_argument(
        "--hud-margin-x", type=float, default=0.05,
        help="Distance of the on-screen counters from the left edge, as a ratio of frame width (0.0-1.0)",
    )
    parser.add_argument(
        "--hud-margin-y", type=float, default=0.06,
        help="Distance of the on-screen counters from the top edge, as a ratio of frame height (0.0-1.0)",
    )
    parser.add_argument(
        "--hud-line-gap", type=float, default=0.06,
        help="Vertical gap between the LV/HV/MC lines, as a ratio of frame height (0.0-1.0)",
    )
    parser.add_argument("--hud-scale", type=float, default=2.0, help="Font scale of the on-screen counters")
    parser.add_argument("--imgsz", type=int, default=640, help="Input resolution for YOLO inference")
    return parser.parse_args()


def process_video(args: argparse.Namespace) -> None:
    device = resolve_device(args.device)
    logger.info("Using device: %s", device)

    class_names = load_class_names(args.classes)
    model = YOLO(args.model)
    model.to(device)

    cap = cv2.VideoCapture(args.video)
    if not cap.isOpened():
        raise FileNotFoundError(f"Could not open video: {args.video}")

    fps = cap.get(cv2.CAP_PROP_FPS)
    time_per_frame = 1 / fps if fps > 0 else 1 / 30
    interval_seconds = args.interval * 60
    roi = load_roi(args.roi_file)

    if not args.no_display:
        cv2.namedWindow("RGB")
        cv2.setMouseCallback("RGB", lambda event, x, y, flags, param: (
            logger.debug("Mouse @ (%d, %d)", x, y) if event == cv2.EVENT_MOUSEMOVE else None
        ))

    prev_positions: dict[int, tuple[float, float, int]] = {}
    counts = IntervalCounts()
    interval_index = 0
    interval_start_seconds = 0.0
    frame_idx = 0
    playback_start = time.perf_counter()

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        frame_idx += 1

        frame = cv2.resize(frame, FRAME_SIZE)
        results = model.track(frame, persist=True, device=device, verbose=False, imgsz=args.imgsz)

        boxes_data = results[0].boxes
        if boxes_data is not None and boxes_data.id is not None:
            boxes = boxes_data.xyxy.cpu().numpy()
            class_ids = boxes_data.cls.cpu().numpy()
            track_ids = boxes_data.id.cpu().numpy()

            for box, class_id, track_id in zip(boxes, class_ids, track_ids):
                x1, y1, x2, y2 = box.astype(int)
                cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
                label = class_names[int(class_id)]
                track_id = int(track_id)

                speed_kmh = None
                if track_id in prev_positions:
                    prev_cx, prev_cy, prev_frame_idx = prev_positions[track_id]
                    distance = np.sqrt((cx - prev_cx) ** 2 + (cy - prev_cy) ** 2)
                    # Use the ACTUAL number of elapsed frames (not always 1), so
                    # speed stays correct even when frames were skipped in
                    # --realtime mode.
                    elapsed_frames = max(frame_idx - prev_frame_idx, 1)
                    elapsed_time = elapsed_frames * time_per_frame
                    if distance > 1:
                        speed_px_per_s = distance / elapsed_time
                        speed_kmh = (speed_px_per_s / PIXEL_TO_METER_SCALE) * 3.6
                prev_positions[track_id] = (cx, cy, frame_idx)

                if label not in VEHICLE_CLASSES:
                    continue
                if cv2.pointPolygonTest(np.array(roi, np.int32), (float(x1), float(y2)), False) < 0:
                    continue

                if speed_kmh is not None and SPEED_MIN_KMH <= speed_kmh <= SPEED_MAX_KMH:
                    if not args.no_display:
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        cv2.circle(frame, (x1, y2), 4, (255, 0, 0), -1)
                        cv2.putText(
                            frame, f"Speed: {speed_kmh:.1f} km/h", (cx, cy - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2,
                        )
                    if track_id not in counts.speed_recorded_ids:
                        counts.speed_records.append((track_id, label, speed_kmh))
                        counts.speed_recorded_ids.add(track_id)

                if label == "car":
                    counts.car_ids.add(track_id)
                elif label == "truck":
                    counts.truck_ids.add(track_id)
                elif label == "bus":
                    counts.bus_ids.add(track_id)
                elif label == "motorcycle":
                    counts.motorcycle_ids.add(track_id)

        # --- Check the interval boundary based on VIDEO time, not wall-clock time ---
        video_time = frame_idx * time_per_frame
        next_interval_boundary = (interval_index + 1) * interval_seconds
        if video_time >= next_interval_boundary:
            save_interval_to_excel(args.output, counts, interval_start_seconds, next_interval_boundary)
            counts = IntervalCounts()
            interval_index += 1
            interval_start_seconds = next_interval_boundary

        if not args.no_display:
            cv2.polylines(frame, [np.array(roi, np.int32)], True, (0, 0, 255), 2)
            hud_x = int(FRAME_SIZE[0] * args.hud_margin_x)
            hud_y = int(FRAME_SIZE[1] * args.hud_margin_y)
            hud_gap = int(FRAME_SIZE[1] * args.hud_line_gap)
            cvzone.putTextRect(frame, f"LV:-{counts.lv}", (hud_x, hud_y), args.hud_scale, 2)
            cvzone.putTextRect(frame, f"HV:-{counts.hv}", (hud_x, hud_y + hud_gap), args.hud_scale, 2)
            cvzone.putTextRect(frame, f"MC:-{counts.mc}", (hud_x, hud_y + 2 * hud_gap), args.hud_scale, 2)
            cv2.imshow("RGB", frame)

            if args.realtime:
                # Preview mode: sync to the video's real speed, skipping frames when behind.
                target_time = frame_idx * time_per_frame
                actual_time = time.perf_counter() - playback_start
                if actual_time < target_time:
                    key = cv2.waitKey(max(1, int((target_time - actual_time) * 1000))) & 0xFF
                else:
                    frames_behind = int((actual_time - target_time) / time_per_frame)
                    for _ in range(frames_behind):
                        if not cap.grab():
                            break
                        frame_idx += 1
                    key = cv2.waitKey(1) & 0xFF
            else:
                # Accuracy mode: every frame is processed, none are skipped.
                key = cv2.waitKey(1) & 0xFF

            if key == ord("q"):
                break

    cap.release()
    if not args.no_display:
        cv2.destroyAllWindows()

    # Flush the final interval if the video ends before a full interval is reached.
    video_end_seconds = frame_idx * time_per_frame
    if not counts.is_empty() or video_end_seconds > interval_start_seconds:
        save_interval_to_excel(args.output, counts, interval_start_seconds, video_end_seconds)


def main() -> None:
    args = parse_args()
    start_time = time.perf_counter()
    process_video(args)
    logger.info("Done. Total processing time: %s", format_hms(time.perf_counter() - start_time))


if __name__ == "__main__":
    main()
